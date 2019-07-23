
import pandas as pd
import openpyxl as px
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from collections import OrderedDict
from mp_common.SpreadSheeter import SpreadSheeter

pd.options.display.max_columns = 100
pd.options.display.max_rows = 100


# MAIN FUNCTION START
def trendy(fileName, tableSelect, sheetSelect, exportXLSXname):
    sheetList = []
    concatList = []

# IMPORT

    for sheet in sheetSelect:
        sheetList.append(sheet)

# POPULATE
    XL = pd.ExcelFile(fileName)
    headersCheck = XL.parse(sheetList[0]).set_index("Name").columns

    if len(tableSelect) == 0:
        headers = XL.parse(sheetList[0]).set_index("Name").columns

    elif any(c in tableSelect for c in headersCheck):
        for f in tableSelect:
            headers = tableSelect

    for a in headers:
        for b in sheetList:
            concatList.append(XL.parse(b)
                                .set_index("Name")[a]
                                .rename(b + " " + a))

    concatSheet = pd.concat(concatList, axis=1, sort=True)

    # print(concatSheet)
    # concatSheet.to_excel(exportXLSXname + ".xlsx",index_label = "Name")

    # OPENPYXL
    wb = px.Workbook()
    ws = wb.active

    for r in dataframe_to_rows(concatSheet, index=True, header=True):
        ws.append(r)
    # COLUMN FORMAT
    for r in range(len(headers)):
        ws.insert_cols(r * (len(sheetList) + 1) + 2, 1)
    # INSERT NAME
    ws.insert_rows(1, 1)
    ws["A2"] = "Name"
    # SUPER-HEADERS
    ws.delete_cols(2, 1)
    for r in range(len(headers)):
        ws[get_column_letter((4*r)+2) + "1"] = headers[r]

    for r in range(len(headers)):
        ws.merge_cells(get_column_letter((4 * r) + 2) + "1" + ":" +
                       get_column_letter((4 * r) + 4) + "1")

    # FINAL FORMATTING
    ws["A2"].font = Font(bold=True)
    for i in range(((len(headers) + 1) * (len(sheetList) + 1))):
        col = get_column_letter(i + 1)
        ws[col + "1"].font = Font(bold=True)
        ws[col + "1"].alignment = Alignment(horizontal="center")
        ws[col + "2"].font = Font(bold=True)
        ws[col + "2"].alignment = Alignment(wrap_text=True)

    ws.delete_rows(3, 1)
    wb.save(exportXLSXname + ".xlsx")
    print(wb)
    pass


# trendy("(Trend)_Talent_Book_2019-06_AFA_All Talent.xlsm",
#        ["Total", "Male", "Female"],
#        ["2019_06_AFA", "2019_05_AFA", "2019_04_AFA"],
#        "cleanupTesting4")

trendy("2019_06_Trend_Local_Test_WB.xlsm",
       ["Total", "Female", "Light Viewers", "Spanish Only"],
       ["2019_05_Appeal", "2019_04_Appeal", "2018_10_Appeal"],
       "ipythonTest4")


# PRINT HEADER OPTIONS
def printHeaders(fileName, sheetName):
    headerPrint = pd.ExcelFile(fileName).parse(sheetName).columns
    print("the following headers can be selected individually to trend %s"
          % headerPrint)
    pass

# printHeaders("(Trend)_Talent_Book_2019-06_AFA_All Talent.xlsm",
#              "2019_06_AFA")
